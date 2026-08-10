# -*- coding: utf-8 -*-
"""26년 2-3월 정산서 전 파일 구조 스캔
- 매직바이트로 실제 파일 형식 판별 (확장자 위장 감지)
- 시트명/행열 크기/헤더 후보 추출
- CSV 인코딩 감지 (utf-8-sig / cp949 / euc-kr)
- 03 수정본 폴더: 배경색 칠해진 셀(=필수 필드 표시) 추출
출력: scan_result.json
"""
import json, os, sys, zipfile, io, re

BASE = "/Users/rachel/workspace/hivemind"
TARGETS = [
    ("02_2월", os.path.join(BASE, "02. 정산서 다운로드 원본", "26년 2월")),
    ("02_3월", os.path.join(BASE, "02. 정산서 다운로드 원본", "26년 3월")),
    ("03_2월수정", os.path.join(BASE, "03. 26년 2월_일부 수정(확장자명 등)")),
]

def sniff(path):
    with open(path, "rb") as f:
        head = f.read(512)
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "ole2_xls"
    if head[:2] == b"PK":
        return "zip_or_xlsx"
    stripped = head.lstrip()
    low = stripped[:200].lower()
    if low.startswith(b"<html") or low.startswith(b"<!doctype") or b"<table" in low or low.startswith(b"<?xml"):
        return "html_or_xml"
    return "text"

def detect_encoding(path):
    raw = open(path, "rb").read()
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            raw.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "unknown"

def cell_str(v):
    if v is None:
        return ""
    s = str(v).strip().replace("\n", " ")
    return s[:40]

_N_XF = 500  # 셀이 참조하는 스타일 인덱스가 범위를 벗어나지 않도록 넉넉히
MINIMAL_STYLES = ("""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<fonts count="1"><font/></fonts><fills count="1"><fill><patternFill patternType="none"/></fill></fills>
<borders count="1"><border/></borders><cellStyleXfs count="1"><xf/></cellStyleXfs>
<cellXfs count="%d">%s</cellXfs></styleSheet>""" % (_N_XF, "<xf/>" * _N_XF)).encode()

def load_wb_tolerant(path):
    """깨진 styles.xml을 가진 비표준 xlsx 폴백 로더"""
    import openpyxl
    try:
        return openpyxl.load_workbook(path, data_only=True), False
    except ValueError:
        src = zipfile.ZipFile(path)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as dst:
            for item in src.infolist():
                if item.filename.endswith("styles.xml"):
                    dst.writestr(item.filename, MINIMAL_STYLES)
                elif not item.is_dir():
                    dst.writestr(item, src.read(item.filename))
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True), True

def scan_xlsx(path, want_colors=False):
    info = {"sheets": []}
    wb, repaired = load_wb_tolerant(path)
    if repaired:
        info["xlsx_styles_repaired"] = True
    for ws in wb.worksheets:
        s = {"name": ws.title, "rows": ws.max_row, "cols": ws.max_column, "head": []}
        for r in ws.iter_rows(min_row=1, max_row=min(6, ws.max_row or 1)):
            s["head"].append([cell_str(c.value) for c in r[:25]])
        if want_colors:
            colored = {}
            for r in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row or 1)):
                for c in r[:30]:
                    f = c.fill
                    if f and f.patternType == "solid":
                        rgb = getattr(f.fgColor, "rgb", None)
                        if rgb and isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                            colored.setdefault(rgb, []).append(f"{c.coordinate}={cell_str(c.value)}")
            s["colored_cells"] = {k: v[:15] for k, v in colored.items()}
        info["sheets"].append(s)
    wb.close()
    return info

def scan_xls(path):
    import xlrd
    info = {"sheets": []}
    wb = xlrd.open_workbook(path)
    for ws in wb.sheets():
        s = {"name": ws.name, "rows": ws.nrows, "cols": ws.ncols, "head": []}
        for ri in range(min(6, ws.nrows)):
            s["head"].append([cell_str(ws.cell_value(ri, ci)) for ci in range(min(25, ws.ncols))])
        info["sheets"].append(s)
    return info

def scan_csv(path):
    import csv
    enc = detect_encoding(path)
    info = {"encoding": enc, "sheets": []}
    if enc == "unknown":
        return info
    with open(path, encoding=enc, newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        delim = "\t" if sample.count("\t") > sample.count(",") else ","
        rows = list(csv.reader(f, delimiter=delim))
    s = {"name": "(csv)", "rows": len(rows), "cols": max((len(r) for r in rows), default=0),
         "delimiter": delim, "head": [[cell_str(v) for v in r[:25]] for r in rows[:6]]}
    info["sheets"].append(s)
    return info

def scan_html(path):
    import pandas as pd
    enc = detect_encoding(path)
    info = {"encoding": enc, "sheets": []}
    try:
        raw = open(path, "rb").read()
        text = raw.decode(enc if enc != "unknown" else "cp949", errors="replace")
        # meta charset 선언이 잘못된 경우 lxml이 죽으므로 제거
        text = re.sub(r"<meta[^>]*charset[^>]*>", "", text, flags=re.I)
        text = re.sub(r"encoding=[\"'][^\"']*[\"']", "", text, count=1)
        tables = pd.read_html(io.StringIO(text))
        for i, df in enumerate(tables):
            info["sheets"].append({
                "name": f"(html_table_{i})", "rows": len(df), "cols": len(df.columns),
                "head": [[cell_str(v) for v in row[:25]] for row in
                         ([list(df.columns)] + df.head(5).values.tolist())]})
    except Exception as e:
        info["error"] = f"html parse fail: {e}"
    return info

def scan_zip(path):
    info = {"zip_members": []}
    with zipfile.ZipFile(path) as z:
        for n in z.namelist():
            try:
                name = n.encode("cp437").decode("cp949")
            except Exception:
                name = n
            info["zip_members"].append(name)
    return info

results = []
for group, folder in TARGETS:
    if not os.path.isdir(folder):
        continue
    for fn in sorted(os.listdir(folder)):
        if fn.startswith(".") or fn.startswith("._"):
            continue
        path = os.path.join(folder, fn)
        if not os.path.isfile(path):
            continue
        ext = os.path.splitext(fn)[1].lower()
        kind = sniff(path)
        entry = {"group": group, "file": fn, "ext": ext, "magic": kind,
                 "size_kb": round(os.path.getsize(path) / 1024, 1)}
        try:
            if kind == "zip_or_xlsx":
                if ext in (".xlsx", ".xlsm"):
                    entry.update(scan_xlsx(path, want_colors=(group == "03_2월수정")))
                    entry["real_type"] = "xlsx"
                else:
                    entry.update(scan_zip(path))
                    entry["real_type"] = "zip"
            elif kind == "ole2_xls":
                entry.update(scan_xls(path))
                entry["real_type"] = "xls(ole2)"
            elif kind == "html_or_xml":
                entry.update(scan_html(path))
                entry["real_type"] = "html위장"
            else:  # text
                entry.update(scan_csv(path))
                entry["real_type"] = "text/csv"
        except Exception as e:
            entry["error"] = repr(e)
        results.append(entry)

out = os.path.join(BASE, "analysis_2026", "scan_result.json")
with open(out, "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=1)

# 콘솔 요약
print(f"{'그룹':<10}{'실형식':<10}{'크기KB':>8}  파일")
for e in results:
    flag = " ⚠️" + e["error"][:40] if "error" in e else ""
    mismatch = " 🔶확장자≠실형식" if e.get("real_type") and not (
        (e["ext"] == ".xlsx" and e["real_type"] == "xlsx") or
        (e["ext"] == ".xls" and e["real_type"] == "xls(ole2)") or
        (e["ext"] == ".csv" and e["real_type"] == "text/csv") or
        (e["ext"] == ".zip" and e["real_type"] == "zip")) else ""
    print(f"{e['group']:<10}{e.get('real_type','?'):<10}{e['size_kb']:>8}  {e['file']}{mismatch}{flag}")
print(f"\n총 {len(results)}개 파일 → {out}")
