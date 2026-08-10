# -*- coding: utf-8 -*-
"""형식 판별 + 관용 로더: 모든 입력을 '시트명 → 2차원 그리드(list of lists)'로 통일
기준 1: 확장자를 믿지 않는다 — 매직바이트로 실제 형식 판별
"""
import io, re, zipfile, csv as csvmod

MINIMAL_STYLES = ("""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<fonts count="1"><font/></fonts><fills count="1"><fill><patternFill patternType="none"/></fill></fills>
<borders count="1"><border/></borders><cellStyleXfs count="1"><xf/></cellStyleXfs>
<cellXfs count="500">%s</cellXfs></styleSheet>""" % ("<xf/>" * 500)).encode()

RIDI_T = re.compile(r'^=T\("(.*)"\)$')  # 리디 =T("...") 수식 래핑


def sniff(path):
    with open(path, "rb") as f:
        head = f.read(512)
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"
    if head[:2] == b"PK":
        return "zip"  # xlsx 또는 zip — 내부로 구분
    low = head.lstrip()[:200].lower()
    if low.startswith(b"<html") or low.startswith(b"<!doctype") or b"<table" in low or low.startswith(b"<?xml"):
        return "html"
    return "text"


def detect_encoding(raw):
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            raw.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "utf-8"


def _clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    m = RIDI_T.match(s)
    if m:
        s = m.group(1).strip()
    return s


def _grid_from_xlsx(stream_or_path):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(stream_or_path, data_only=True, read_only=True)
    except ValueError:  # 비표준 styles.xml (북큐브·애니툰) → 교체 후 재시도
        src = zipfile.ZipFile(stream_or_path)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as dst:
            for item in src.infolist():
                if item.filename.endswith("styles.xml"):
                    dst.writestr(item.filename, MINIMAL_STYLES)
                elif not item.is_dir():
                    dst.writestr(item, src.read(item.filename))
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)
    sheets = {}
    for ws in wb.worksheets:
        try:
            ws.reset_dimensions()  # 거짓 dimension 메타데이터 방어 (카카오 등)
        except AttributeError:
            pass
        sheets[ws.title] = [[_clean(c.value) if not isinstance(c.value, (int, float)) else c.value
                             for c in row] for row in ws.iter_rows()]
    wb.close()
    return sheets


def _grid_from_xls(path):
    import xlrd
    wb = xlrd.open_workbook(path)
    sheets = {}
    for ws in wb.sheets():
        grid = []
        for ri in range(ws.nrows):
            row = []
            for ci in range(ws.ncols):
                v = ws.cell_value(ri, ci)
                row.append(v if isinstance(v, (int, float)) else _clean(v))
            grid.append(row)
        sheets[ws.name] = grid
    return sheets


def _cell(v):
    if isinstance(v, float) and v != v:  # NaN
        return ""
    return v if isinstance(v, (int, float)) else _clean(v)


def _grid_from_html(path):
    """pandas는 <thead>를 df.columns로 흡수하므로 헤더를 그리드 맨 위에 복원한다"""
    import pandas as pd
    raw = open(path, "rb").read()
    enc = detect_encoding(raw)
    text = raw.decode(enc, errors="replace")
    text = re.sub(r"<meta[^>]*charset[^>]*>", "", text, flags=re.I)
    text = re.sub(r"encoding=[\"'][^\"']*[\"']", "", text, count=1)
    # 여는 <tr> 누락 보정 (피플앤스토리형 깨진 HTML)
    text = re.sub(r"</tr>\s*<td", "</tr><tr><td", text, flags=re.I)
    tables = pd.read_html(io.StringIO(text))
    sheets = {}
    for i, df in enumerate(tables):
        grid = []
        cols = df.columns
        headerless = isinstance(cols, pd.RangeIndex) or all(
            str(c).strip().lstrip("-").isdigit() for c in cols)
        if not headerless:  # 흡수된 thead 복원 (정수 컬럼명 = 헤더 없음 → 복원 생략)
            levels = cols.nlevels
            for lv in range(levels):
                line = []
                for c in cols:
                    name = c[lv] if levels > 1 else c
                    s = "" if name is None else str(name)
                    if s.startswith("Unnamed:") or s == "nan":
                        s = ""
                    line.append(s.strip())
                grid.append(line)
        for row in df.itertuples(index=False):
            grid.append([_cell(v) for v in row])
        sheets[f"table_{i}"] = grid
    return sheets


def _grid_from_csv(path_or_text):
    if isinstance(path_or_text, str) and "\n" not in path_or_text:
        raw = open(path_or_text, "rb").read()
        text = raw.decode(detect_encoding(raw))
    else:
        text = path_or_text
    sample = text[:4096]
    delim = "\t" if sample.count("\t") > sample.count(",") else ","
    rows = list(csvmod.reader(io.StringIO(text), delimiter=delim))
    return {"csv": [[_clean(v) for v in r] for r in rows]}


def load_grids(path):
    """경로 → {시트명: grid}. 형식은 내용으로 판별."""
    kind = sniff(path)
    if kind == "zip":
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            if any(n.startswith("xl/") or n == "[Content_Types].xml" for n in names):
                return _grid_from_xlsx(path), "xlsx"
            # 일반 zip (리디): 내부 첫 데이터 파일을 재귀 처리
            inner = [n for n in names if not n.endswith("/")][0]
            data = z.read(inner)
            if data[:2] == b"PK":
                return _grid_from_xlsx(io.BytesIO(data)), "zip>xlsx"
            text = data.decode(detect_encoding(data))
            return _grid_from_csv(text), "zip>csv"
    if kind == "xls":
        return _grid_from_xls(path), "xls"
    if kind == "html":
        return _grid_from_html(path), "html"
    return _grid_from_csv(path), "csv"


def flatten_header(grid, row_start, row_end, max_col=None):
    """다단 헤더 평탄화: 빈 칸은 왼쪽 값 상속(병합 그룹), 행 간 '/'로 결합"""
    max_col = max_col or max(len(r) for r in grid[row_start - 1:row_end])
    merged = []
    for r in range(row_start - 1, row_end):
        row = [str(v).strip() if v != "" else "" for v in grid[r]] + [""] * max_col
        carry = ""
        line = []
        for i in range(max_col):
            if row[i]:
                carry = row[i]
            line.append(carry)
        merged.append(line)
    names = []
    for i in range(max_col):
        parts = []
        for m in merged:
            p = m[i]
            if p and (not parts or parts[-1] != p):
                parts.append(p)
        names.append("/".join(parts))
    return names
